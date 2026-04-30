"""Per-company unpaid-commission tracker.

Snapshots are taken inside the comparison route — every time an agent uploads
a commission file, we group the `only_production` (unpaid) customers from
that compare result by their paying company and upsert one
`UnpaidSnapshot` row per (user, commission_upload, company).

Idempotency: the unique constraint on (user_id, commission_upload_id, company_name)
+ ON CONFLICT DO UPDATE means re-uploading the same file replaces the prior
snapshot — no duplicates. Combined with the existing replace-on-upload semantics
for FileUpload (commission_pairing memory), this gives clean MoM trends.
"""
import io
import logging
import uuid
from collections import defaultdict
from datetime import datetime
from decimal import Decimal

from sqlalchemy import select, desc, and_
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.unpaid_snapshot import UnpaidSnapshot
from app.models.unpaid_dismissal import UnpaidDismissal
from app.models.upload import FileUpload
from app.models.company_contact import CompanyContact
from app.models.commission_rate import CommissionRate
from app.models.user import User
from app.services import email_service

# Category keyword hints for picking the right commission rate when an agent
# has multiple rates per insurance company (e.g. one for גמל, one for ביטוח).
_GEMEL_HINTS = ["גמל", "השתלמות", "תגמולים", "חיסכון פלוס"]
_INSURANCE_HINTS = ["חיים", "בריאות", "סיעוד", "חסכון פרט", "פיננסי", "משכנתא"]

logger = logging.getLogger(__name__)


def _period_label(uploaded_at: datetime | None) -> str:
    if not uploaded_at:
        uploaded_at = datetime.utcnow()
    return uploaded_at.strftime("%Y-%m")


def _name_matches(a: str | None, b: str | None) -> bool:
    """Substring fuzzy match (case-insensitive). Same idea as the frontend."""
    if not a or not b:
        return False
    al, bl = a.lower(), b.lower()
    return al in bl or bl in al


def _category_hints_for(product: dict) -> list[str] | None:
    """Pick a category hint set based on whether the product looks like
    gemel (has accumulation) or insurance (has premium)."""
    if (product.get("accumulation") or 0) > 0:
        return _GEMEL_HINTS
    if (product.get("premium") or 0) > 0:
        return _INSURANCE_HINTS
    return None


def _find_rate(product: dict, rates: list[CommissionRate]) -> float | None:
    """Mirror of the frontend findRate() in ComparisonDashboard.vue.

    Find a commission rate matching the product's company. When multiple match,
    prefer the one matching the product's category (gemel/insurance).
    """
    if not rates:
        return None
    candidates: list[str] = []
    for f in (product.get("company"), product.get("company_full")):
        if f and f.strip():
            candidates.append(f.strip())
    if not candidates:
        return None

    matches: list[CommissionRate] = []
    for cn in candidates:
        cl = cn.lower()
        first_word = cl.split()[0] if cl.split() else ""
        for r in rates:
            rn = (r.company_name or "").lower()
            if not rn:
                continue
            if r.company_name == cn or cl in rn or rn in cl or (
                len(first_word) > 2 and rn.startswith(first_word)
            ):
                if r not in matches:
                    matches.append(r)

    if not matches:
        return None
    if len(matches) == 1:
        return float(matches[0].rate)

    hints = _category_hints_for(product)
    if hints:
        preferred = next(
            (r for r in matches if any(h in (r.company_name or "") for h in hints)),
            None,
        )
        if preferred:
            return float(preferred.rate)
    return float(matches[0].rate)


def _calc_expected_commission(product: dict, rate: float | None) -> float | None:
    """Python port of `frontend/src/utils/commissionCalc.js#calcExpectedCommission`."""
    if not rate:
        return None
    accum = product.get("accumulation") or 0
    if accum:
        return float(accum) * rate / 12
    balance = product.get("balance") or 0
    if balance:
        return float(balance) * rate / 12
    premium = product.get("premium") or 0
    if premium:
        return float(premium) * rate * 100
    return None


def _product_matches_commission(
    product: dict,
    commission_companies: list[str],
) -> bool:
    """Match the product against any commission company on EITHER its short
    `company` field OR its full `company_full` field.

    Production parsers populate `company` with the short brand (e.g. אינטרגמל,
    a sub-brand) while `company_full` is the legal entity (e.g. "מור גמל ופנסיה
    בע\"מ"). Without checking both we miss the very thing the user uploaded for.
    """
    short = (product.get("company") or "").strip()
    full = (product.get("company_full") or "").strip()
    return any(
        _name_matches(short, c) or _name_matches(full, c)
        for c in commission_companies
    )


def _resolve_company(
    production_products: list[dict],
    commission_companies: list[str],
) -> str | None:
    """Pick which commission company this customer belongs to in this snapshot.

    Returns the canonical commission company name (e.g. "מור") — NOT the
    production product's variant — so the monitor groups by the company we
    actually uploaded for.
    """
    if not commission_companies:
        return None
    for p in production_products or []:
        for c in commission_companies:
            short = (p.get("company") or "").strip()
            full = (p.get("company_full") or "").strip()
            if _name_matches(short, c) or _name_matches(full, c):
                return c
    return None


async def snapshot_unpaid_after_compare(
    db: AsyncSession,
    user_id: uuid.UUID,
    commission_upload_id: uuid.UUID,
    comparison_result: dict,
) -> int:
    """Snapshot per-company unpaid customers from a comparison result.

    Returns the number of (company) rows upserted.
    """
    # Resolve the upload to derive period_label from uploaded_at
    upload = (await db.execute(
        select(FileUpload).where(FileUpload.id == commission_upload_id)
    )).scalar_one_or_none()
    if not upload:
        return 0

    period = _period_label(upload.uploaded_at)
    snapshot_at = datetime.utcnow()

    # Load commission rates for this user — used to compute expected commission
    # per unpaid product, mirroring `totalUnpaidCharge` in the dashboard.
    rates = (await db.execute(
        select(CommissionRate).where(CommissionRate.user_id == user_id)
    )).scalars().all()

    # Restrict to customers tied to the commission company this upload covers.
    # Without this, every only_production customer (incl. ones whose products
    # are from totally different companies) gets dragged in — we'd be saying
    # "מור didn't pay you for an הראל customer", which is nonsense.
    commission_companies = (
        comparison_result.get("commission_company_sources")
        or ([comparison_result.get("commission_company_source")] if comparison_result.get("commission_company_source") else [])
        or []
    )
    commission_companies = [c for c in commission_companies if c]
    if not commission_companies:
        # Nothing to attribute the unpaid set to — bail without writing anything.
        return 0

    # Group only_production (unpaid) customers by their paying company,
    # filtered to ONLY customers with at least one product from a commission
    # company source. Mirrors the frontend `relevantCustomers` logic.
    by_company: dict[str, list[dict]] = defaultdict(list)
    for c in comparison_result.get("customers", []) or []:
        if c.get("match_status") != "only_production":
            continue
        company = _resolve_company(c.get("production_products") or [], commission_companies)
        if not company:
            continue
        # Keep only the production products that belong to this commission
        # company — same scoping the dashboard does.
        relevant_products = [
            p for p in (c.get("production_products") or [])
            if _product_matches_commission(p, commission_companies)
        ]
        if not relevant_products:
            continue
        first = (c.get("first_name") or "").strip()
        last = (c.get("last_name") or "").strip()
        full_name = (f"{first} {last}").strip() or None

        products = []
        cust_raw = 0.0          # sum of premium||accumulation (exposure)
        cust_expected = 0.0     # sum of expected commission (the actual ₪ unpaid)
        for p in relevant_products:
            premium = float(p.get("premium") or 0)
            accum = float(p.get("accumulation") or 0)
            raw = premium if premium > 0 else accum
            rate = _find_rate(p, rates)
            expected = _calc_expected_commission(p, rate) or 0.0
            cust_raw += raw
            cust_expected += expected
            products.append({
                "product": p.get("product"),
                "company": p.get("company"),
                "premium": premium,
                "accumulation": accum,
                "value": raw,
                "expected_commission": expected,
                "rate": rate,
                "policy_number": p.get("policy_number"),
            })

        # Skip customers with no value to chase. Same gemel filter the dashboard
        # uses (effectiveUnpaidCustomers): customers whose relevant products
        # are all 0/null get dropped — there's nothing to recover from them.
        if cust_raw == 0:
            continue

        by_company[company].append({
            "id_number": c.get("id_number"),
            "name": full_name,
            "premium": cust_expected if cust_expected > 0 else cust_raw,
            "expected_commission": cust_expected,
            "raw_value": cust_raw,
            "products": products,
        })

    if not by_company:
        # Still wipe any old snapshots that pointed at this upload (e.g. a prior
        # compare had unpaid for some company; the new compare has none)
        await db.execute(
            UnpaidSnapshot.__table__.delete().where(
                and_(
                    UnpaidSnapshot.user_id == user_id,
                    UnpaidSnapshot.commission_upload_id == commission_upload_id,
                )
            )
        )
        await db.commit()
        return 0

    # Upsert one row per company. ON CONFLICT replaces count/amount/customers
    # but PRESERVES email_sent_at so a previously-mailed snapshot stays marked.
    upserted = 0
    for company, customers in by_company.items():
        # Sum of expected unpaid commission across all customers in this group.
        # Falls back to raw exposure when no rates are configured — matches the
        # dashboard's `totalUnpaidCharge` (`expectedTotal || rawTotal`).
        expected_sum = sum(float(c.get("expected_commission") or 0) for c in customers)
        raw_sum = sum(float(c.get("raw_value") or 0) for c in customers)
        amount = expected_sum if expected_sum > 0 else raw_sum
        total_amount = Decimal(str(amount)).quantize(Decimal("0.01"))
        stmt = pg_insert(UnpaidSnapshot).values(
            user_id=user_id,
            commission_upload_id=commission_upload_id,
            company_name=company[:100],
            period_label=period,
            unpaid_count=len(customers),
            total_amount=total_amount,
            unpaid_customers_json=customers,
            snapshot_date=snapshot_at,
            created_at=snapshot_at,
        )
        stmt = stmt.on_conflict_do_update(
            constraint="uq_unpaid_snapshots_user_upload_company",
            set_={
                "period_label": stmt.excluded.period_label,
                "unpaid_count": stmt.excluded.unpaid_count,
                "total_amount": stmt.excluded.total_amount,
                "unpaid_customers_json": stmt.excluded.unpaid_customers_json,
                "snapshot_date": stmt.excluded.snapshot_date,
            },
        )
        await db.execute(stmt)
        upserted += 1

    # Also drop snapshots for companies no longer in the result for this upload
    seen_companies = list(by_company.keys())
    await db.execute(
        UnpaidSnapshot.__table__.delete().where(
            and_(
                UnpaidSnapshot.user_id == user_id,
                UnpaidSnapshot.commission_upload_id == commission_upload_id,
                ~UnpaidSnapshot.company_name.in_(seen_companies),
            )
        )
    )
    await db.commit()
    return upserted


async def _dismissed_ids_for_user(
    db: AsyncSession,
    user_id: uuid.UUID,
) -> dict[str, set[str]]:
    """Map of company_name → set of dismissed customer_id_numbers."""
    rows = (await db.execute(
        select(UnpaidDismissal).where(UnpaidDismissal.user_id == user_id)
    )).scalars().all()
    out: dict[str, set[str]] = {}
    for r in rows:
        out.setdefault(r.company_name, set()).add(r.customer_id_number)
    return out


async def current_by_company(db: AsyncSession, user_id: uuid.UUID) -> list[dict]:
    """Aggregate per company across all snapshots.

    For each company the user has commission data for, returns one row showing
    UNIQUE unpaid customers across every snapshot that company has — covering
    all months the user has uploaded for. A customer counted as unpaid in
    multiple periods only counts once. The amount sums the *latest* observed
    expected commission per unique customer.

    `snapshot_id` returned is the most recent snapshot's id (used as the target
    for send-email actions and drill-in).
    """
    rows = (await db.execute(
        select(UnpaidSnapshot, FileUpload.uploaded_at)
        .join(FileUpload, FileUpload.id == UnpaidSnapshot.commission_upload_id)
        .where(UnpaidSnapshot.user_id == user_id)
        .order_by(desc(FileUpload.uploaded_at))
    )).all()

    dismissed = await _dismissed_ids_for_user(db, user_id)

    # Group by company → walk newest-first, aggregating unique customers
    by_company: dict[str, dict] = {}
    for snap, uploaded_at in rows:
        company = snap.company_name
        company_dismissed = dismissed.get(company, set())
        if company not in by_company:
            by_company[company] = {
                "snapshot_id": str(snap.id),  # most recent snapshot id
                "company": company,
                "period": snap.period_label,
                "uploaded_at": uploaded_at.isoformat() if uploaded_at else None,
                "email_sent_at": snap.email_sent_at.isoformat() if snap.email_sent_at else None,
                "email_sent_to": snap.email_sent_to,
                "_customers": {},   # id_number -> customer dict (newest wins)
                "_periods": set(),
            }
        bucket = by_company[company]
        bucket["_periods"].add(snap.period_label)
        for c in snap.unpaid_customers_json or []:
            cid = (c.get("id_number") or "").strip()
            if not cid or cid in company_dismissed:
                continue
            # Newest-first iteration — first time we see this customer is the
            # most recent snapshot they appear in. Skip later (older) duplicates.
            if cid in bucket["_customers"]:
                continue
            bucket["_customers"][cid] = c

    out: list[dict] = []
    for company, bucket in by_company.items():
        customers = list(bucket["_customers"].values())
        # Skip companies whose entire customer set has been dismissed.
        if not customers:
            continue
        amount = sum(
            float(c.get("expected_commission") or c.get("premium") or 0)
            for c in customers
        )
        period_range = sorted(bucket["_periods"])
        if not period_range:
            period_label = bucket["period"]
        elif len(period_range) == 1:
            period_label = period_range[0]
        else:
            period_label = f"{period_range[0]} → {period_range[-1]}"
        out.append({
            "snapshot_id": bucket["snapshot_id"],
            "company": company,
            "period": period_label,
            "count": len(customers),
            "amount": float(amount),
            "email_sent_at": bucket["email_sent_at"],
            "email_sent_to": bucket["email_sent_to"],
            "uploaded_at": bucket["uploaded_at"],
            "customers_preview": customers[:5],
        })

    out.sort(key=lambda r: r["amount"], reverse=True)
    return out


async def trend(
    db: AsyncSession,
    user_id: uuid.UUID,
    company: str | None = None,
    months: int = 12,
) -> list[dict]:
    """Per-period rows, optionally filtered by company. Used for the line chart.

    Returns one row per (period, company) — the frontend pivots into series.
    For a given (period, company), if multiple uploads exist we sum the counts
    and amounts (e.g. one March commission file for גמל and another for ביטוח
    both showing מור unpaid → both contribute to the מור / 2026-03 point).
    """
    q = select(UnpaidSnapshot).where(UnpaidSnapshot.user_id == user_id)
    if company:
        q = q.where(UnpaidSnapshot.company_name == company)
    snapshots = (await db.execute(q)).scalars().all()

    # Aggregate by (period, company)
    agg: dict[tuple[str, str], dict] = {}
    for s in snapshots:
        key = (s.period_label, s.company_name)
        if key not in agg:
            agg[key] = {
                "period": s.period_label,
                "company": s.company_name,
                "count": 0,
                "amount": 0.0,
            }
        agg[key]["count"] += int(s.unpaid_count or 0)
        agg[key]["amount"] += float(s.total_amount or 0)

    rows = list(agg.values())
    # Keep last N months
    periods = sorted({r["period"] for r in rows}, reverse=True)[:months]
    rows = [r for r in rows if r["period"] in periods]
    rows.sort(key=lambda r: (r["period"], r["company"]))
    return rows


def _build_unpaid_xlsx(
    company: str,
    period: str,
    customers: list[dict],
) -> bytes:
    """Build an XLSX file with the unpaid customer list for one company."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "לא_שולם"
    ws.sheet_view.rightToLeft = True

    headers = ["ת.ז", "שם", "מוצרים", "פרמיה", "צבירה", "מספר פוליסה"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="F57C00")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for c in customers:
        products = c.get("products") or []
        product_names = " | ".join(p.get("product") or "" for p in products if p.get("product"))
        total_premium = sum(float(p.get("premium") or 0) for p in products) or float(c.get("premium") or 0)
        total_accum = sum(float(p.get("accumulation") or 0) for p in products)
        policies = " | ".join(p.get("policy_number") or "" for p in products if p.get("policy_number"))
        ws.append([
            c.get("id_number") or "",
            c.get("name") or "",
            product_names,
            round(total_premium, 2),
            round(total_accum, 2),
            policies,
        ])

    # Column widths
    widths = [16, 28, 40, 14, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def send_company_email(
    db: AsyncSession,
    user_id: uuid.UUID,
    snapshot_id: uuid.UUID,
) -> dict:
    """Build an XLSX of the unpaid customers and email it to the company contact.

    Raises ValueError("no_contact") if the user hasn't configured an email
    for that company yet — the API layer maps that to a 404 with a friendly
    "אין מייל מוגדר ל-X" message.
    """
    snap = (await db.execute(
        select(UnpaidSnapshot).where(
            and_(UnpaidSnapshot.id == snapshot_id, UnpaidSnapshot.user_id == user_id)
        )
    )).scalar_one_or_none()
    if not snap:
        raise ValueError("not_found")

    contact = (await db.execute(
        select(CompanyContact).where(
            and_(
                CompanyContact.user_id == user_id,
                CompanyContact.company_name == snap.company_name,
            )
        )
    )).scalar_one_or_none()
    if not contact or not contact.email:
        raise ValueError("no_contact")

    user = (await db.execute(select(User).where(User.id == user_id))).scalar_one()
    agent_name = user.full_name or user.email

    customers = snap.unpaid_customers_json or []
    xlsx_bytes = _build_unpaid_xlsx(snap.company_name, snap.period_label, customers)

    await email_service.send_unpaid_company_email(
        to_email=contact.email,
        agent_name=agent_name,
        company_name=snap.company_name,
        period_label=snap.period_label,
        customers=customers,
        xlsx_bytes=xlsx_bytes,
    )

    snap.email_sent_at = datetime.utcnow()
    snap.email_sent_to = contact.email
    await db.commit()

    return {
        "ok": True,
        "sent_to": contact.email,
        "sent_at": snap.email_sent_at.isoformat(),
        "count": len(customers),
    }


async def dismiss_customer(
    db: AsyncSession,
    user_id: uuid.UUID,
    company: str,
    customer_id_number: str,
) -> bool:
    """Hide a single customer from the unpaid view for one company.

    Idempotent — re-dismissing an already-dismissed customer is a no-op. The
    snapshot data is left intact so the trend chart history is preserved.
    """
    if not company or not customer_id_number:
        return False
    existing = (await db.execute(
        select(UnpaidDismissal).where(
            and_(
                UnpaidDismissal.user_id == user_id,
                UnpaidDismissal.company_name == company,
                UnpaidDismissal.customer_id_number == customer_id_number,
            )
        )
    )).scalar_one_or_none()
    if existing:
        return True
    db.add(UnpaidDismissal(
        user_id=user_id,
        company_name=company,
        customer_id_number=customer_id_number,
    ))
    await db.commit()
    return True


async def dismiss_company(
    db: AsyncSession,
    user_id: uuid.UUID,
    company: str,
) -> int:
    """Hide every currently-visible customer for one company.

    Walks the company's existing snapshots, collects unique customer ids, and
    bulk-inserts dismissal rows. Returns the number of customers dismissed.
    """
    if not company:
        return 0
    rows = (await db.execute(
        select(UnpaidSnapshot).where(
            and_(
                UnpaidSnapshot.user_id == user_id,
                UnpaidSnapshot.company_name == company,
            )
        )
    )).scalars().all()

    already = {
        d.customer_id_number
        for d in (await db.execute(
            select(UnpaidDismissal).where(
                and_(
                    UnpaidDismissal.user_id == user_id,
                    UnpaidDismissal.company_name == company,
                )
            )
        )).scalars().all()
    }

    to_add: set[str] = set()
    for s in rows:
        for c in s.unpaid_customers_json or []:
            cid = (c.get("id_number") or "").strip()
            if cid and cid not in already and cid not in to_add:
                to_add.add(cid)

    for cid in to_add:
        db.add(UnpaidDismissal(
            user_id=user_id,
            company_name=company,
            customer_id_number=cid,
        ))
    if to_add:
        await db.commit()
    return len(to_add)


async def restore_customer(
    db: AsyncSession,
    user_id: uuid.UUID,
    company: str,
    customer_id_number: str | None = None,
) -> int:
    """Undo dismissal. If `customer_id_number` is None, restores the whole company."""
    q = UnpaidDismissal.__table__.delete().where(
        and_(
            UnpaidDismissal.user_id == user_id,
            UnpaidDismissal.company_name == company,
        )
    )
    if customer_id_number:
        q = q.where(UnpaidDismissal.customer_id_number == customer_id_number)
    result = await db.execute(q)
    await db.commit()
    return result.rowcount or 0


async def get_customers(
    db: AsyncSession,
    user_id: uuid.UUID,
    snapshot_id: uuid.UUID,
) -> list[dict] | None:
    """Drill-in: full unpaid customer list for the company that owns this snapshot.

    To match the aggregated `current_by_company` view, this returns the UNIQUE
    customers across every snapshot for the same company — newest-first, so
    the customer's most recent unpaid record wins. That way the drill list
    matches the count shown on the row.
    """
    snap = (await db.execute(
        select(UnpaidSnapshot).where(
            and_(UnpaidSnapshot.id == snapshot_id, UnpaidSnapshot.user_id == user_id)
        )
    )).scalar_one_or_none()
    if not snap:
        return None

    rows = (await db.execute(
        select(UnpaidSnapshot, FileUpload.uploaded_at)
        .join(FileUpload, FileUpload.id == UnpaidSnapshot.commission_upload_id)
        .where(
            and_(
                UnpaidSnapshot.user_id == user_id,
                UnpaidSnapshot.company_name == snap.company_name,
            )
        )
        .order_by(desc(FileUpload.uploaded_at))
    )).all()

    dismissed = await _dismissed_ids_for_user(db, user_id)
    company_dismissed = dismissed.get(snap.company_name, set())

    seen: dict[str, dict] = {}
    for s, _uploaded_at in rows:
        for c in s.unpaid_customers_json or []:
            cid = (c.get("id_number") or "").strip()
            if not cid or cid in seen or cid in company_dismissed:
                continue
            seen[cid] = c
    return list(seen.values())
