import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.user import User
from app.models.debt import Debt
from app.models.company_contact import CompanyContact
from app.api.deps import get_paid_user as get_current_user
from app.services.debt_service import get_debts_summary
from app.services.email_service import send_debt_email

router = APIRouter()


class DebtOut(BaseModel):
    id: str
    company_name: str
    category: str
    customer_id_number: str
    customer_name: str
    product: str | None
    policy_number: str | None
    expected_amount: float
    premium: float | None
    accumulation: float | None
    status: str
    status_changed_at: str | None
    last_emailed_at: str | None
    email_count: int
    created_at: str


class StatusUpdate(BaseModel):
    status: str  # open, paid, disputed, cancelled


class BulkStatusUpdate(BaseModel):
    ids: list[str]
    status: str


class BulkEmailRequest(BaseModel):
    company_name: str


def _debt_to_out(d: Debt) -> DebtOut:
    return DebtOut(
        id=str(d.id),
        company_name=d.company_name,
        category=d.category,
        customer_id_number=d.customer_id_number,
        customer_name=d.customer_name,
        product=d.product,
        policy_number=d.policy_number,
        expected_amount=float(d.expected_amount),
        premium=float(d.premium) if d.premium else None,
        accumulation=float(d.accumulation) if d.accumulation else None,
        status=d.status,
        status_changed_at=d.status_changed_at.isoformat() if d.status_changed_at else None,
        last_emailed_at=d.last_emailed_at.isoformat() if d.last_emailed_at else None,
        email_count=d.email_count,
        created_at=d.created_at.isoformat(),
    )


@router.get("", response_model=list[DebtOut])
async def list_debts(
    status: str | None = Query(default=None),
    company: str | None = Query(default=None),
    category: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(Debt).where(Debt.user_id == user.id)
    if status:
        q = q.where(Debt.status == status)
    if company:
        q = q.where(Debt.company_name == company)
    if category:
        q = q.where(Debt.category == category)
    q = q.order_by(desc(Debt.expected_amount))

    result = await db.execute(q)
    return [_debt_to_out(d) for d in result.scalars().all()]


@router.get("/summary")
async def debts_summary(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return await get_debts_summary(db, user.id)


@router.patch("/{debt_id}/status")
async def update_debt_status(
    debt_id: str,
    data: StatusUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if data.status not in ("open", "paid", "disputed", "cancelled"):
        raise HTTPException(status_code=400, detail="סטטוס לא חוקי")

    result = await db.execute(
        select(Debt).where(Debt.id == uuid.UUID(debt_id), Debt.user_id == user.id)
    )
    debt = result.scalar_one_or_none()
    if not debt:
        raise HTTPException(status_code=404, detail="חוב לא נמצא")

    debt.status = data.status
    debt.status_changed_at = datetime.utcnow()
    await db.commit()
    return _debt_to_out(debt)


@router.patch("/bulk-status")
async def bulk_update_status(
    data: BulkStatusUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if data.status not in ("open", "paid", "disputed", "cancelled"):
        raise HTTPException(status_code=400, detail="סטטוס לא חוקי")

    updated = 0
    for did in data.ids:
        result = await db.execute(
            select(Debt).where(Debt.id == uuid.UUID(did), Debt.user_id == user.id)
        )
        debt = result.scalar_one_or_none()
        if debt:
            debt.status = data.status
            debt.status_changed_at = datetime.utcnow()
            updated += 1

    await db.commit()
    return {"updated": updated}


@router.post("/send-email")
async def send_company_debt_email(
    data: BulkEmailRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Send debt summary email to a company's contact."""
    # Find company contact
    contact_q = await db.execute(
        select(CompanyContact).where(
            CompanyContact.user_id == user.id,
            CompanyContact.company_name.ilike(f"%{data.company_name}%"),
        )
    )
    contact = contact_q.scalar_one_or_none()
    if not contact or not contact.email:
        raise HTTPException(status_code=404, detail="לא נמצא אימייל לחברה זו. הוסף אימייל בלשונית 'אימיילים לחברות'")

    # Get open debts for this company
    debts_q = await db.execute(
        select(Debt).where(
            Debt.user_id == user.id,
            Debt.company_name == data.company_name,
            Debt.status == "open",
        ).order_by(desc(Debt.expected_amount))
    )
    debts = debts_q.scalars().all()
    if not debts:
        raise HTTPException(status_code=400, detail="אין חובות פתוחים לחברה זו")

    agent_name = user.full_name or user.email
    debts_list = [
        {
            "customer_name": d.customer_name,
            "customer_id": d.customer_id_number,
            "product": d.product or "—",
            "policy_number": d.policy_number or "—",
            "amount": float(d.expected_amount),
        }
        for d in debts
    ]

    try:
        await send_debt_email(
            to_email=contact.email,
            company_name=data.company_name,
            agent_name=agent_name,
            debts_list=debts_list,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאה בשליחת אימייל: {str(e)}")

    # Update email tracking
    for d in debts:
        d.last_emailed_at = datetime.utcnow()
        d.email_count = (d.email_count or 0) + 1
    await db.commit()

    return {"ok": True, "sent_to": contact.email, "debts_count": len(debts)}


@router.get("/export/excel")
async def export_debts_excel(
    status: str | None = Query(default="open"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export debts as Excel file."""
    import openpyxl

    q = select(Debt).where(Debt.user_id == user.id)
    if status:
        q = q.where(Debt.status == status)
    q = q.order_by(Debt.company_name, desc(Debt.expected_amount))

    result = await db.execute(q)
    debts = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "סיכום חובות"
    ws.sheet_view.rightToLeft = True

    headers = ["חברה", "קטגוריה", "שם לקוח", "ת.ז", "מוצר", "מספר פוליסה", "סכום צפוי", "פרמיה", "צבירה", "סטטוס"]
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")

    status_map = {"open": "פתוח", "paid": "שולם", "disputed": "במחלוקת", "cancelled": "בוטל"}

    for d in debts:
        ws.append([
            d.company_name,
            "גמל והשתלמות" if d.category == "gemel_hishtalmut" else "ביטוח",
            d.customer_name,
            d.customer_id_number,
            d.product or "—",
            d.policy_number or "—",
            float(d.expected_amount),
            float(d.premium) if d.premium else None,
            float(d.accumulation) if d.accumulation else None,
            status_map.get(d.status, d.status),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=debts_summary.xlsx"},
    )
